import subprocess
import sys
import os

# --- ติดตั้งไลบรารีแบบ auto ถ้ายังไม่มี ---
def install_if_missing(package_name):
    try:
        __import__(package_name)
    except ImportError:
        print(f"Installing {package_name}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name, "--target", os.path.dirname(__file__)])
        print(f"{package_name} installed successfully!")

install_if_missing("openpyxl")
install_if_missing("pyperclip")

# --- import ไลบรารีหลังติดตั้ง ---
import pyperclip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- ฟังก์ชันหลัก ---
def highlight_missing_fields(filepath, save_path=None):
    wb = load_workbook(filepath)
    ws = wb.active

    last_row = ws.max_row
    last_col = ws.max_column
    output_col = last_col + 1  # Output column for result (Missing Fields)

    ws.cell(row=1, column=output_col).value = "Missing Fields"

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="11FF00", end_color="11FF00", fill_type="solid")
    clear_fill = PatternFill(fill_type=None)

    for i in range(2, last_row + 1):
        missing_list = ""
        is_complete = True

        for j in range(1, last_col + 1):
            cell = ws.cell(row=i, column=j)
            header = ws.cell(row=1, column=j).value
            cell.fill = clear_fill

            if cell.value is None or str(cell.value).strip() == "":
                cell.fill = red_fill
                missing_list += header + ", "
                is_complete = False

        result_cell = ws.cell(row=i, column=output_col)
        result_cell.fill = clear_fill

        if is_complete:
            result_cell.value = "ALL FIELDS ARE FILLED"
            result_cell.fill = green_fill
        else:
            result_cell.value = missing_list[:-2]

    # กำหนดชื่อไฟล์ใหม่
    base, ext = os.path.splitext(os.path.basename(filepath))
    new_filename = base + "_CheckedFields" + ext

    if save_path is None:
        save_filepath = filepath
    else:
        save_filepath = os.path.join(save_path, new_filename)

    wb.save(save_filepath)
    print(f"✅ Saved file to: {save_filepath}")

    # คัดลอก path ไป clipboard
    pyperclip.copy(save_path)
    print("📋 File path copied to clipboard!")

# --- เรียกใช้งานตรงนี้ ---
highlight_missing_fields(
    r"C:\ExampleFolder\ExampleSubFolder\ExampleFile.xlsx",
    save_path=r"C:\ExampleFolder\ExampleSubFolder"
)